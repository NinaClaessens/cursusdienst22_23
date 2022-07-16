import csv
from datetime import datetime
import openpyxl

match = [(4, 14), (5, 15), (6, 9), (7, 10), (8, 11)]
cur_year = str(datetime.now().year)[-2:]


def get_years_back(nb_years):
    return str(int(cur_year) - nb_years)


all_data = dict()
titles = ["Vakken"]

wb = openpyxl.Workbook()
ws = wb.active

with open('sale_articles_0_2021-2022.csv') as f:
    reader = csv.reader(f, delimiter=',')
    for row in reader:
        ws.append(row)

first = True
ibarcode, ivakcode, ivaknaam = 0, 0, 0
for row in ws.iter_rows(values_only=True):
    if first:
        for i, cel in enumerate(row):
            if cel == "Barcode":
                ibarcode = i
            elif cel == "Code":
                ivakcode = i
            elif cel == "Vak":
                ivaknaam = i
            else:
                titles.append(cel)
        first = False
    else:
        barcode = row[ibarcode]
        if barcode in all_data:
            new_vak = (row[ivakcode], row[ivaknaam])
            all_data[barcode][0].append(new_vak)
        else:
            vak = [(row[ivakcode], row[ivaknaam])]
            data = [vak]
            for i, cel in enumerate(row):
                if i not in [ibarcode, ivakcode, ivaknaam]:
                    data.append(cel)
            all_data[barcode] = data

print(titles)
print(all_data)

wb = openpyxl.load_workbook('cursussen.xlsx')
ws = wb.active

for row in ws.iter_rows(min_row=2):
    if row[0].value is not None and row[1].value is not None and type(row[1].value) is int:
        barcode = '978' + get_years_back(1) + cur_year + str(row[1].value)
        if barcode in all_data:
            data = all_data[barcode]
            for idata, irow in match:
                store = row[irow].value
                if idata == 8:
                    row[irow].value = 1 if data[idata] == '1' else 0
                else:
                    row[irow].value = int(data[idata])
        else:
            print('nay', barcode)
wb.save('cursussen_adapted.xlsx')

